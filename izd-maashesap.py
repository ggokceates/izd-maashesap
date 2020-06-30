import sqlite3
import openpyxl

con = sqlite3.connect("calisanlar.db")
cursor = con.cursor()


# sabitler (TİS)
ykc = 111.92
ymk_g = 17.91
biry = 279.80


def tablo_olustur():
    cursor.execute("CREATE TABLE IF NOT EXISTS kullanicilar ("
                   "sicil TEXT,"
                   "isim TEXT,"
                   "soyisim TEXT,"
                   "agi FLOAT,"
                   "uct FLOAT,"
                   "cck INT,"
                   "sai FLOAT)"
                   )
    con.commit()


def deger_ekle():
    cursor.execute("INSERT INTO kullanicilar "
                   "VALUES (?,?,?,?,?,?,?)", (sicil, isim, soyisim, agi, uct, cck, sai))
    con.commit()
    con.close()


def dbden_veri_al():
    sicil = input("Lütfen sicilinizi giriniz: ")
    cursor.execute("SELECT * FROM kullanicilar WHERE sicil =?", (sicil,))
    con.commit()
    data = cursor.fetchone()
    isim = data[1]
    soyisim = data[2]
    agi = data[3]
    uct = data[4]
    cck = data[5]
    sai = data[6]
    return isim, soyisim, agi, uct, cck, sai


def dbden_veri_al_forxlc():
    cursor.execute("SELECT * FROM kullanicilar WHERE sicil =?", (sicil,))
    con.commit()
    data = cursor.fetchone()
    isim = data[1]
    soyisim = data[2]
    agi = data[3]
    uct = data[4]
    cck = data[5]
    sai = data[6]
    return isim, soyisim, agi, uct, cck, sai


def sabit_veri_al():
    agi = input("AGİ Giriniz: ")  #float
    uct = input("Brüt Ücret: ")   #float
    cck = input("Çocuk Sayısı: ") or "0"
    sai = input("Sendika Aidatı: ")  #float
    return agi, uct, cck, sai


def aylik_veri_al():
    ngun = int(input("Çalışma Gün Sayısı: "))
    ht = input("Hafta Tatili Sayısı: ")

    gm = input("Gece Mesai (Saat): ")  # float
    nm = input("Normal Mesai (Saat):")  # float
    uby = input("3Y Bayram Mesai (Saat): ")  # float
    iby = input("2Y Bayram Mesai (Saat): ")  # float
    htc = input("Hafta Tatili Çalışması: ")
    rpr = input("Rapor: ")
    try:
        yi = int(input("Yıllık İzin: "))
    except ValueError:
        yi = 0
    mi = input("Mazeret İzni: ")

    ia = input("İkramiye Gün (Aralık ayı için 22, diğer aylar için 18 giriniz): ")

    return ngun, ht, gm, nm, uby, iby, htc, rpr, yi, mi, ia


def excelden_veri_al():
    sicil = ws.cell(row=rw, column=1).value
    ngun = ws.cell(row=rw, column=4).value
    ht = ws.cell(row=rw, column=5).value
    gm = ws.cell(row=rw, column=6).value
    nm = ws.cell(row=rw, column=7).value
    uby = ws.cell(row=rw, column=8).value
    iby = ws.cell(row=rw, column=9).value
    htc = ws.cell(row=rw, column=10).value
    rpr = ws.cell(row=rw, column=11).value
    yi = ws.cell(row=rw, column=12).value
    mi = ws.cell(row=rw, column=13).value
    ia = ws.cell(row=rw, column=14).value
    return sicil, ngun, ht, gm, nm, uby, iby, htc, rpr, yi, mi, ia


def std_maas_hsp():
    try:
        ccky = 54.28 * int(cck)         #TİS

    except ValueError:
        ccky = 0

    ngun_u = float(uct) * int(ngun) # normal gün hesabı (zorunlu veri)
    ht_u = float(uct) * int(ht)  # hafta tatili hesabı (zorunlu veri)
    ymk = float(ymk_g) * (ngun+yi)   # yemek parası hesabı (TİS)

    try:
        gm_u = float(gm) * (float(uct) / 7.5 * 0.2)  # gece mesaisi hesabı (TİS)
    except ValueError:
        gm_u = 0

    try:
        nm_u = float(nm) * (float(uct) / 7.5 * 1.6)  # normal mesai hesabı (TİS)
    except ValueError:
        nm_u = 0

    try:
        uby_u = float(uby) * (float(uct) / 7.5 * 3)  # 3y bayram mesai hesabı (TİS)
    except ValueError:
        uby_u = 0

    try:
        iby_u = float(iby) * (float(uct) / 7.5 * 2)  # 2y bayram mesai hesabı (TİS)
    except ValueError:
        iby_u = 0

    try:
        htc_u = int(htc) * (float(uct) * 2)  # hafta tatili çalışması (TİS)
    except ValueError:
        htc_u = 0

    try:
        ikr_u = float(uct) * int(ia)  # ikramiye (TİS)
    except ValueError:
        ikr_u = 0

    try:
        yi_u = float(uct) * int(yi)
    except ValueError:
        yi_u = 0

    try:
        rpr_u = float(uct) * int(rpr)
    except ValueError:
        rpr_u = 0

    try:
        mi_u = float(uct) * int(mi)
    except ValueError:
        mi_u = 0



    sgk_m = ngun_u+ht_u+gm_u+nm_u+uby_u+iby_u+htc_u+ikr_u+yi_u+rpr_u+mi_u+ykc+float(ymk)+biry+int(ccky)-float(sai)

    isc_p = sgk_m * 0.14  # sgk işçi primi
    isz_p = sgk_m * 0.01  # işsizlik primi
    gv_m = sgk_m - (isc_p + isz_p)  # gelir vergisi matrahı
    gv = gv_m * (int(vd) * 0.01)  # gelir vergisi
    dv = sgk_m * 0.00759  # damga vergisi

    net = float(sgk_m)-(float(isc_p)+float(isz_p)+float(gv)+float(dv))+float(agi)

    return ngun_u, ht_u, ykc, ymk, biry, ccky, sgk_m, isc_p, isz_p, gv_m, gv, dv, gm_u, \
           nm_u, uby_u, iby_u, htc_u, ikr_u, yi_u, rpr_u, mi_u, net


def veri_dokum():
    print("Normal Gün: ", format(ngun_u, '.2f'))
    print("Hafta Tatili: ", ht_u)

    print("Yakacak: ", ykc,"\t                                  \t Gece Mesaisi: ", format(gm_u, '.2f'))
    print("Yemek Ücreti: ", ymk,"\t                             \t Normal Mesai: ", format(nm_u, '.2f'))
    print("Bir. Yardım: ", biry,"\t                             \t 3Y Bayram Mesaisi: ", format(uby_u, '.2f'))
    print("Çocuk Yardımı: ", ccky,"\t                               \t 2Y Bayram Mesaisi: ", format(iby_u, '.2f'))
    print("SGK Matrahı: ", format(sgk_m, '.2f'),"\t                              \t Hafta Tatili Çalışması: ", htc_u)
    print("SGK İşçi Primi: ", format(isc_p, '.2f'),"\t                          \t Yıllık İzin: ", yi_u)
    print("İşsizlik Primi: ", format(isz_p, '.2f'),"\t                              \t Rapor: ", rpr_u)
    print("Gelir Vergisi Matrahı: ", format(gv_m, '.2f'),"\t                  \t Mazeret İzni: ", mi_u)
    print("Gelir Vergisi: ", format(gv, '.2f'),"\t                               \t İkramiye :", ikr_u)
    print("Damga Vergisi: ", format(dv, '.2f'))

    #print("Gece Mesaisi: ", format(gm_u, '.2f'))
    #print("Normal Mesai: ", format(nm_u, '.2f'))
    #print("3Y Bayram Mesaisi: ", format(uby_u, '.2f'))
    #print("2Y Bayram Mesaisi: ", format(iby_u, '.2f'))
    #print("Hafta Tatili Çalışması: ", htc_u)
    #print("Yıllık İzin: ", yi_u)
    #print("Rapor: ", rpr_u)
    #print("Mazeret İzni: ", mi_u)
    #print("İkramiye :", ikr_u)

    print("-------------------------------------")
    print("NET Ödenecek: ", format(net, '.2f'))
    print("-------------------------------------")
    print("\n")




while True:

    sec = input("1 - Kayıtlı Kullanıcı \n"
                "2 - Yeni Kullanıcı \n"
                "3 - Misafir Girişi \n"
                "4 - Excel Dosyasından Veri Al \n"
                "Lütfen Seçim Yapınız: ")


    if sec == "1":

        print("***Ondalık sayıları . ile giriniz!***")

        isim, soyisim, agi, uct, cck, sai = dbden_veri_al()

        ngun, ht, gm, nm, uby, iby, htc, rpr, yi, mi, ia = aylik_veri_al()

        vd = input("Lütfen Vergi Diliminizi Giriniz (Bilmiyorsanız 20'den hesaplanır): ") or "20"
        print("\n")

        ngun_u, ht_u, ykc, ymk, biry, ccky, sgk_m, isc_p, isz_p, gv_m, \
        gv, dv, gm_u, nm_u, uby_u, iby_u, htc_u, ikr_u, yi_u, rpr_u, mi_u, net = std_maas_hsp()

        print(isim + " " + soyisim)
        print("----------------------------------")

        veri_dokum()


    if sec == "2":

        print("***Ondalık sayıları . ile giriniz!***")

        tablo_olustur()

        sicil = input("Sicil No: ")
        isim = input("İsim: ")
        soyisim = input("Soyisim: ")

        agi, uct, cck, sai = sabit_veri_al()

        deger_ekle()

        ngun, ht, gm, nm, uby, iby, htc, rpr, yi, mi, ia = aylik_veri_al()

        vd = input("Lütfen Vergi Diliminizi Giriniz (Bilmiyorsanız 20'den hesaplanır): ") or "20"
        print("\n")

        ngun_u, ht_u, ykc, ymk, biry, ccky, sgk_m, isc_p, isz_p, gv_m, \
        gv, dv, gm_u, nm_u, uby_u, iby_u, htc_u, ikr_u, yi_u, rpr_u, mi_u, net = std_maas_hsp()

        print(isim + " " + soyisim)
        print("----------------------------------")

        veri_dokum()


    if sec == "3":

        print("***Ondalık sayıları . ile giriniz!***")

        agi, uct, cck, sai = sabit_veri_al()

        ngun, ht, gm, nm, uby, iby, htc, rpr, yi, mi, ia = aylik_veri_al()

        vd = input("Lütfen Vergi Diliminizi Giriniz (Bilmiyorsanız 20'den hesaplanır): ") or "20"
        print("\n")

        ngun_u, ht_u, ykc, ymk, biry, ccky, sgk_m, isc_p, isz_p, gv_m, \
        gv, dv, gm_u, nm_u, uby_u, iby_u, htc_u, ikr_u, yi_u, rpr_u, mi_u, net = std_maas_hsp()

        print("----------------------------------")

        veri_dokum()


    if sec == "4":

        print("***Ondalık sayıları . ile giriniz!***")

        from openpyxl import load_workbook
        dosya_ismi = input("Lütfen dosya ismini giriniz: ")
        wb = load_workbook(dosya_ismi+".xlsx")

        rw = 2
        ws_ismi = input("Lütfen Çalışma Sayfası ismini giriniz: ")
        ws = wb[ws_ismi]

        vd = input("Lütfen Vergi Diliminizi Giriniz (Bilmiyorsanız 20'den hesaplanır): ") or "20"
        print("\n")

        while True:

            if ws.cell(row = rw, column = 1).value == None:
                break

            sicil, ngun, ht, gm, nm, uby, iby, htc, rpr, yi, mi, ia = excelden_veri_al()

            isim, soyisim, agi, uct, cck, sai = dbden_veri_al_forxlc()

            ngun_u, ht_u, ykc, ymk, biry, ccky, sgk_m, isc_p, isz_p, gv_m, \
            gv, dv, gm_u, nm_u, uby_u, iby_u, htc_u, ikr_u, yi_u, rpr_u, mi_u, net = std_maas_hsp()

            print(isim+" "+soyisim)
            print("----------------------------------")

            veri_dokum()

            rw += 1


    else:
        ex = input("Çıkmak için bir tuşa basınız")
    break
